from dash import html, dcc, Input, Output, State, no_update, ctx, dash_table
import dash_bootstrap_components as dbc
import pandas as pd
from datetime import datetime
import base64
import io
import os
import traceback
from pathlib import Path
import openpyxl

from app.pages.base import LimsDashApp  
from app.core.database import SessionLocal
from app.schema.objects import Order, Sample #, PanelMaster
from app.core.mapping import FACILITY_MAPPING, SPECIAL_MAPPING, get_full_mapping_for_panel


def create_registration_layout():
    # 실제로는 Facility 테이블을 조회하는 것이 좋으나 기존 문맥 유지를 위해 MAPPING 사용
    facility_opts = [{"label": f"[{code}] {data['facility']} ({data['team']})", "value": code} for code, data in FACILITY_MAPPING.items()]
    # facility_opts = [{"label": f"[{code}] {data['facility']} ({data['team']})", "value": code} for code, data in SPECIAL_MAPPING.items()]


    return html.Div([
        html.H3("📥 검사 의뢰서 신규 접수", className="fw-bold mb-4 text-secondary"),
        html.P("의뢰 기관과 검사 종류를 지정한 후 엑셀 의뢰서를 업로드해 주세요.", className="text-muted"),
        
        dbc.Card([
            dbc.CardBody([
                html.H5("1. 접수 기본 정보 및 의뢰자 정보", className="fw-bold text-primary mb-3"),
                
                dbc.Row([
                    dbc.Col([
                        html.Label("1-1. 의뢰 기관 (필수)", className="fw-bold text-danger small mb-2"),
                        dcc.Dropdown(id="reg-facility-select", options=facility_opts, placeholder="기관 코드 선택...", className="shadow-sm mb-3")
                    ], width=6),
                    dbc.Col([
                        html.Label("1-2. 검사 종류 (필수)", className="fw-bold text-danger small mb-2"),
                        dcc.Dropdown(id="reg-panel-select", options=[], placeholder="DB에서 활성화된 검사 종류를 불러오는 중...", className="shadow-sm mb-3")
                    ], width=6)
                ]),
                
                html.Div("💡 아래 의뢰자 정보를 먼저 입력하고 양식을 다운로드하면, 해당 정보가 엑셀에 자동으로 채워진 상태로 생성됩니다.", className="text-muted small mb-2"),
                dbc.Row([
                    dbc.Col([
                        html.Label("의뢰자 성명", className="fw-bold text-secondary small mb-1"),
                        dbc.Input(id="reg-client-name", placeholder="예: 홍길동", className="shadow-sm mb-3")
                    ], width=4),
                    dbc.Col([
                        html.Label("연락처", className="fw-bold text-secondary small mb-1"),
                        dbc.Input(id="reg-client-phone", placeholder="예: 010-1234-5678", className="shadow-sm mb-3")
                    ], width=4),
                    dbc.Col([
                        html.Label("이메일", className="fw-bold text-secondary small mb-1"),
                        dbc.Input(id="reg-client-email", type="email", placeholder="예: email@example.com", className="shadow-sm mb-3")
                    ], width=4)
                ], className="p-3 bg-light rounded border mb-3"),

                dbc.Row([
                    dbc.Col([
                        dbc.Button("📥 DB 기준 맞춤 양식 다운로드", id="btn-download-template", outline=True, color="primary", className="w-100 fw-bold shadow-sm"),
                        dcc.Download(id="download-template-file")
                    ], width=12)
                ], className="mb-3"),
                
                html.Div(id="template-preview-container", className="mb-4 p-3 bg-white border border-info rounded shadow-sm"),
                html.Hr(className="my-4 border-secondary"),

                html.H5("2. 작성된 의뢰서 업로드", className="fw-bold text-primary mb-3"),
                dcc.Upload(
                    id='upload-excel-data',
                    children=html.Div([
                        '📂 작성 완료된 의뢰서(Excel)를 여기에 ', html.B('드래그 앤 드롭', className="text-primary"), ' 하거나 ',
                        html.A('클릭하여 선택하세요.', className="text-decoration-underline text-primary fw-bold", style={'cursor': 'pointer'})
                    ]),
                    style={'width': '100%', 'height': '80px', 'lineHeight': '80px', 'borderWidth': '2px', 'borderStyle': 'dashed', 'borderColor': '#18BC9C', 'borderRadius': '10px', 'textAlign': 'center', 'marginBottom': '10px', 'backgroundColor': '#f8fffb'},
                    multiple=False
                ),
                html.Div(id="upload-filename-display", className="text-success fw-bold mb-3 small ms-1"),

                html.Div(id="parsed-data-container", style={"display": "none"}, children=[
                    html.H5("3. 추출된 데이터 확인", className="fw-bold text-primary mb-3 mt-4"),
                    dcc.Store(id="parsed-order-store"),
                    dbc.Alert(id="order-info-alert", color="success", className="shadow-sm py-2"),
                    dash_table.DataTable(
                        id="parsed-sample-table", columns=[], data=[],
                        style_table={'overflowX': 'auto', 'border': '1px solid #dee2e6'},
                        style_cell={'textAlign': 'center', 'padding': '10px', 'fontFamily': 'sans-serif', 'minWidth': '100px', 'whiteSpace': 'normal'},
                        style_header={'backgroundColor': '#f8f9fa', 'fontWeight': 'bold'},
                        page_size=10
                    ),
                    dbc.Button("🚀 추출된 데이터 LIMS에 최종 등록 (접수 대기)", id="btn-save-new", color="primary", size="lg", className="w-100 mt-4 fw-bold shadow-sm")
                ]),
                
                html.Div(id="save-new-message", className="mt-3")
            ])
        ], className="border-0 shadow-sm rounded-4")
    ])

# ==========================================
# [2] 콜백 로직
# ==========================================
def register_registration_callbacks(dash_app):
    
    # 🚀 [MODIFIED] 페이지 로드 시 DB에서 Panel Master 목록을 조회하여 드롭다운 구성
    @dash_app.callback(Output("reg-panel-select", "options"), Input("reg-panel-select", "id"))
    def load_panel_options(_):
        db = SessionLocal()
        try:
            panels = db.query(PanelMaster).filter(PanelMaster.is_active == 1).all()
            return [{"label": f"{p.panel_name} ({p.panel_code})", "value": p.panel_code} for p in panels]
        finally:
            db.close()
            
    # 🚀 [MODIFIED] 선택된 패널의 DB 설정값(request_template_name)에 따라 엑셀 템플릿 로드
    @dash_app.callback(Output("template-preview-container", "children"), Input("reg-panel-select", "value"))
    def update_template_preview(panel_code):
        if not panel_code: return html.Div("👆 검사 종류(1-2)를 선택하시면 의뢰서 미리보기가 나타납니다.", className="text-muted small text-center py-2")
        
        db = SessionLocal()
        try:
            panel = db.query(PanelMaster).filter(PanelMaster.panel_code == panel_code).first()
            if not panel: return html.Div("⚠️ 등록되지 않은 검사 항목입니다.", className="text-danger small")
            template_name = panel.request_template_name
        finally:
            db.close()
            
        excel_file_path = os.path.join(Path(__file__).parent.parent / "templates" / "requests", f"{template_name}.xlsx")
        if not os.path.exists(excel_file_path): return html.Div(f"⚠️ 매핑된 양식 파일이 서버에 없습니다: {template_name}.xlsx", className="text-danger small")
        
        try:
            df_raw = pd.read_excel(excel_file_path, header=None)
            header_idx = next((r for r in range(min(25, len(df_raw))) if "patientid" in "".join(df_raw.iloc[r].astype(str).tolist()).replace(" ", "").lower()), 15)
            df_preview = pd.read_excel(excel_file_path, header=header_idx).dropna(axis=1, how='all').head(5)
            df_preview.columns = [str(c) if not str(c).startswith('Unnamed') else '' for c in df_preview.columns]
            return html.Div([html.H6(f"[{panel.panel_name}] 지정 양식 미리보기", className="fw-bold text-primary mb-2"), dash_table.DataTable(columns=[{"name": str(i), "id": str(i)} for i in df_preview.columns if str(i)], data=df_preview.to_dict('records'), style_table={'overflowX': 'auto', 'border': '1px solid #dee2e6'}, style_cell={'textAlign': 'center', 'padding': '10px', 'fontSize': '12px'}, style_header={'backgroundColor': '#e9ecef', 'fontWeight': 'bold'})])
        except Exception as e: return html.Div(f"⚠️ 미리보기 오류: {e}", className="text-danger small")

    @dash_app.callback(
        Output("download-template-file", "data"), 
        Input("btn-download-template", "n_clicks"), 
        [State("reg-panel-select", "value"),
         State("reg-client-name", "value"),
         State("reg-client-phone", "value"),
         State("reg-client-email", "value")], 
        prevent_initial_call=True
    )
    def download_excel_template(n_clicks, panel_code, client_name, client_phone, client_email):
        if not panel_code: return no_update
        
        db = SessionLocal()
        try:
            panel = db.query(PanelMaster).filter(PanelMaster.panel_code == panel_code).first()
            template_name = panel.request_template_name if panel else "default_request"
        finally:
            db.close()
            
        excel_file_path = os.path.join(Path(__file__).parent.parent / "templates" / "requests", f"{template_name}.xlsx")
        
        if not os.path.exists(excel_file_path): 
            return dcc.send_data_frame(pd.DataFrame().to_excel, f"{template_name}.xlsx", index=False)
            
        try:
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=1, max_row=15):
                for cell in row:
                    if not cell.value: continue
                    val = str(cell.value).replace(" ", "").lower()
                    if "실무자성명" in val:
                        ws.cell(row=cell.row, column=cell.column + 1).value = client_name or ""
                    elif "연락처" in val:
                        ws.cell(row=cell.row, column=cell.column + 1).value = client_phone or ""
                    elif "e-mail" in val or "email" in val:
                        ws.cell(row=cell.row, column=cell.column + 1).value = client_email or ""

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            prefix = f"{client_name.replace(' ', '')}_" if client_name else ""
            download_name = f"{prefix}_{template_name}.xlsx"
            return dcc.send_bytes(out.getvalue(), download_name)
            
        except Exception as e:
            print(f"템플릿 수정 중 에러 발생: {e}")
            return dcc.send_file(excel_file_path)
    
    @dash_app.callback(
        [Output("parsed-data-container", "style"), Output("upload-filename-display", "children"), Output("order-info-alert", "children"), Output("parsed-sample-table", "columns"), Output("parsed-sample-table", "data"),
         Output("reg-client-name", "value"), Output("reg-client-phone", "value"), Output("reg-client-email", "value")],
        Input("upload-excel-data", "contents"), [State("upload-excel-data", "filename"), State("reg-facility-select", "value"), State("reg-panel-select", "value")], prevent_initial_call=True
    )
    def parse_and_preview_excel(contents, filename, facility_code, panel_code):
        if not contents: return {"display": "none"}, "", "", [], [], no_update, no_update, no_update
        if not facility_code or not panel_code: return {"display": "none"}, dbc.Alert("🚨 업로드 전 1-1(기관)과 1-2(검사 종류)를 선택해주세요.", color="danger", className="py-2 mb-0"), "", [], [], no_update, no_update, no_update

        try:
            content_type, content_string = contents.split(',')
            decoded = base64.b64decode(content_string)
            fac_info = FACILITY_MAPPING.get(facility_code, {"facility": "Unknown", "team": "Unknown"})
            
            df_raw = pd.read_excel(io.BytesIO(decoded), header=None)
            
            client_name, client_phone, client_email = "", "", ""
            for r in range(2, min(12, len(df_raw))): 
                row_vals = df_raw.iloc[r].dropna().astype(str).str.strip().tolist()
                row_nospace = [x.replace(" ", "").lower() for x in row_vals]
                
                if "실무자성명" in row_nospace:
                    idx_name = row_nospace.index("실무자성명")
                    if idx_name + 1 < len(row_vals): client_name = row_vals[idx_name + 1]
                    if "연락처" in row_nospace:
                        idx_phone = row_nospace.index("연락처")
                        if idx_phone + 1 < len(row_vals): client_phone = row_vals[idx_phone + 1]
                    if r + 1 < len(df_raw):
                        next_row_vals = df_raw.iloc[r+1].dropna().astype(str).str.strip().tolist()
                        next_row_nospace = [x.replace(" ", "").lower() for x in next_row_vals]
                        if "e-mail" in next_row_nospace or "email" in next_row_nospace:
                            idx_email = next_row_nospace.index("e-mail") if "e-mail" in next_row_nospace else next_row_nospace.index("email")
                            if idx_email + 1 < len(next_row_vals): client_email = next_row_vals[idx_email + 1]

            header_idx = next((r for r in range(5, min(25, len(df_raw))) if "patientid" in "".join(df_raw.iloc[r].astype(str).tolist()).replace(" ", "").lower()), 15)
            df_samples = pd.read_excel(io.BytesIO(decoded), header=header_idx).dropna(axis=1, how='all')
            df_samples.columns = [str(c) if not str(c).startswith('Unnamed') else f"col_{i}" for i, c in enumerate(df_samples.columns)]
            
            pid_col = next((col for col in df_samples.columns if "patient id" in col.lower() or "번호" in col.lower()), None)
            if pid_col:
                df_samples = df_samples.dropna(subset=[pid_col])
                df_samples = df_samples[df_samples[pid_col].astype(str).str.lower() != 'ex']
            
            df_samples = df_samples.fillna("")
            dynamic_cols = [{"name": c, "id": c} for c in df_samples.columns if "col_" not in c]
            table_data = df_samples.to_dict('records')
            
            alert_ui = html.Div([
                html.Strong("📂 타겟: "), html.Span(f"[{facility_code}] {fac_info['facility']} ({fac_info['team']}) / {panel_code}", className="me-3 text-primary"),
                html.Strong("📊 추출: "), html.Span(f"{len(table_data)}건")
            ])
            return {"display": "block"}, f"✅ {filename} 파싱 성공!", alert_ui, dynamic_cols, table_data, client_name, client_phone, client_email
            
        except Exception as e:
            print(traceback.format_exc())
            return {"display": "none"}, dbc.Alert(f"🚨 파싱 오류: {e}", color="danger", className="py-2 mb-0"), "", [], [], no_update, no_update, no_update

    # 🚀 최종 DB 저장
    @dash_app.callback(
        Output("save-new-message", "children"),
        Input("btn-save-new", "n_clicks"),
        [State("parsed-sample-table", "data"), 
         State("reg-facility-select", "value"), 
         State("reg-panel-select", "value"),
         State("reg-client-name", "value"),
         State("reg-client-phone", "value"),
         State("reg-client-email", "value")], 
        prevent_initial_call=True
    )
    def save_final_data_to_db(n_clicks, sample_data, facility_code, panel_code, client_name, client_phone, client_email):
        if not n_clicks: return no_update
        if not sample_data or not facility_code or not panel_code: 
            return dbc.Alert("⚠️ 필수 정보 누락 또는 파싱된 데이터가 없습니다.", color="warning")
        
        final_client_name = client_name if client_name else "-"
        final_client_phone = client_phone if client_phone else "-"
        final_client_email = client_email if client_email else "-"

        def get_fuzzy_val(raw_dict, target_key):
            if target_key in raw_dict: return raw_dict[target_key]
            clean_target = str(target_key).lower().replace(" ", "").replace("\n", "")
            for k, v in raw_dict.items():
                if clean_target in str(k).lower().replace(" ", "").replace("\n", ""):
                    return v
            return None

        db = SessionLocal()
        try:
            # 🚀 [MODIFIED] PanelMaster에서 핵심 제어 변수 가져오기
            panel_info = db.query(PanelMaster).filter(PanelMaster.panel_code == panel_code).first()
            if not panel_info:
                return dbc.Alert("🚨 DB에 해당 패널 정보가 존재하지 않습니다.", color="danger")
                
            today_str = datetime.now().strftime("%y%m%d")
            fac_info = FACILITY_MAPPING.get(facility_code, {"facility": "Unknown", "team": "Unknown"})
            
            total_orders_today = db.query(Order).filter(Order.order_id.like(f"%{today_str}%")).count()
            batch_seq = str(total_orders_today + 1).zfill(2)

            new_order_id = f"GCX-{facility_code}-{today_str}-{batch_seq}"
            
            new_order = Order(
                order_id=new_order_id, 
                facility=fac_info["facility"], 
                client_team=fac_info["team"],
                client_name=final_client_name,      
                client_email=final_client_email,    
                client_phone=final_client_phone,    
                reception_date=datetime.utcnow().date(), 
                reception_type="일반",
                sales_unit_price=0
            )
            db.add(new_order)
            db.flush() 
            
            mapping_rule = get_full_mapping_for_panel(panel_code)
            
            sample_seq_counter = 0 
            db_insert_count = 0     
            
            for raw_row in sample_data:
                base_data, extra_metadata = {}, {}
                for excel_col, map_info in mapping_rule.items():
                    val = get_fuzzy_val(raw_row, excel_col) 
                    clean_val = str(val).strip() if val is not None and str(val).strip() not in ["nan", "NaT", ""] else None
                    if map_info["is_extra"]: extra_metadata[map_info["db_col"]] = clean_val
                    else: base_data[map_info["db_col"]] = clean_val

                if not base_data.get("sample_name"):
                    fallback_val = (get_fuzzy_val(raw_row, "Patient ID") or 
                                    get_fuzzy_val(raw_row, "Sample ID") or 
                                    get_fuzzy_val(raw_row, "환자번호") or 
                                    get_fuzzy_val(raw_row, "검체번호") or
                                    get_fuzzy_val(raw_row, "Patient ID/ Sample ID"))
                    if fallback_val and str(fallback_val).strip() not in ["", "nan", "NaT"]:
                        base_data["sample_name"] = str(fallback_val).strip()

                if not base_data.get("sample_name"): 
                    continue

                sample_seq_counter += 1
                sample_seq = str(sample_seq_counter).zfill(3) 

                # 🚀 [MODIFIED] 하드코딩된 'if TSO500' 대신 PanelMaster의 설정값을 읽어 DNA/RNA 생성 여부 결정
                target_na = panel_info.target_nucleic_acid.upper()
                types_to_create = []
                if "BOTH" in target_na or "DNA/RNA" in target_na:
                    types_to_create = ["DNA", "RNA"]
                elif "RNA" in target_na:
                    types_to_create = ["RNA"]
                else:
                    types_to_create = ["DNA"]

                for na_type in types_to_create:
                    internal_sample_id = f"ACC-{today_str}-{batch_seq}-{sample_seq}-{na_type}"
                    
                    new_sample = Sample(
                        order_pk=new_order.id,               
                        order_id=new_order.order_id,         
                        sample_id=internal_sample_id, 
                        target_panel=panel_code, 
                        nucleic_acid_type=na_type,       
                        current_status="접수 대기",          
                        sample_name=base_data["sample_name"], 
                        cancer_type=base_data.get("cancer_type"), 
                        specimen=base_data.get("specimen"),
                        project_name=base_data.get("sample_group") or "Default_Project", 
                        pairing_info=base_data.get("pairing_info"), 
                        outside_id_1=base_data.get("outside_id_1"),
                        issue_comment=base_data.get("issue_comment"), 
                        sample_received="대기중",
                        visual_inspection="대기중",
                        panel_metadata=extra_metadata
                    )
                    db.add(new_sample)
                    db_insert_count += 1

            if db_insert_count == 0:
                db.rollback()
                return dbc.Alert("🚨 엑셀에서 유효한 검체 정보(Patient ID)를 찾지 못했습니다. 매핑 양식을 확인해 주세요.", color="danger")

            db.commit() 
            return dbc.Alert(f"🎉 성공! 의뢰자[{final_client_name}]님의 원본 검체 {sample_seq_counter}건 (총 {db_insert_count}개 튜브) 등록이 완료되었습니다.", color="success")
        
        except Exception as e:
            db.rollback()
            print(traceback.format_exc())
            return dbc.Alert(f"🚨 DB 저장 오류: {e}", color="danger")
        finally: 
            db.close()

def create_registration_app(requests_pathname_prefix: str):
    lims = LimsDashApp(__name__, requests_pathname_prefix)
    lims.set_content(create_registration_layout)
    app = lims.get_app()
    register_registration_callbacks(app)
    return app