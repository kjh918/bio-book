import openpyxl

def scan_all_theme_colors(file_path):
    wb = openpyxl.load_workbook(file_path)
    themes = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill.start_color.type == 'theme':
                    idx = cell.fill.start_color.theme
                    if idx not in themes:
                        themes[idx] = cell.coordinate
    
    print("--- Detected Theme Indices in Template ---")
    for idx, coord in themes.items():
        print(f"Theme Index: {idx} (Found first at {coord})")
		
import openpyxl

def scan_font_theme_colors(file_path):
    wb = openpyxl.load_workbook(file_path)
    font_themes = {}
    
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # 폰트 컬러가 'theme' 타입인 경우 추출
                if cell.font and cell.font.color and cell.font.color.type == 'theme':
                    idx = cell.font.color.theme
                    if idx not in font_themes:
                        font_themes[idx] = cell.coordinate
                        
    print("--- Detected [FONT] Theme Indices ---")
    for idx, coord in font_themes.items():
        print(f"Theme Index: {idx} (Found first at {coord})")

# 실행
scan_font_theme_colors("data/templete.xlsx")
scan_all_theme_colors("data/templete.xlsx")