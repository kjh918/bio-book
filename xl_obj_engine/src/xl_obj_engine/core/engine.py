import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, Color, PatternFill
from openpyxl.cell.rich_text import CellRichText
from .models import (
    WorkbookModel, SheetModel, CellModel, StyleModel, 
    BorderModel, BorderSideModel, MergedCellModel, DimensionModel, FillModel
)

class ExcelObjectEngine:
    THEME_RGB_MAP = {
        0: "F2F2F2", 2: "D0CECE", 4: "B4C6E7", 7: "FFE699", 9: "A9D08E",
    }

    @staticmethod
    def _to_openpyxl_color(color_info):
        if color_info is None or str(color_info) in ["None", "00000000"]:
            return None
        if hasattr(color_info, 'rgb'):
            raw_rgb = color_info.rgb
            if color_info.theme in ExcelObjectEngine.THEME_RGB_MAP:
                raw_rgb = ExcelObjectEngine.THEME_RGB_MAP[color_info.theme]
        else:
            raw_rgb = str(color_info)
        clean_hex = raw_rgb.replace("#", "").upper()
        if not clean_hex or clean_hex == "00000000": return None
        if len(clean_hex) == 6: clean_hex = f"FF{clean_hex}"
        return Color(rgb=clean_hex) if len(clean_hex) == 8 else None

    @staticmethod
    def _safe_get_rgb(color_obj):
        if color_obj is None: return "00000000"
        try:
            if color_obj.type == 'rgb' and isinstance(color_obj.rgb, str):
                return color_obj.rgb
        except: pass
        return "00000000"

    @staticmethod
    def read_to_model(file_path: str) -> WorkbookModel:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheet_models = []
        for ws in wb.worksheets:
            merged_list = [
                MergedCellModel(
                    range_string=str(m), 
                    master_coordinate=f"{get_column_letter(range_boundaries(str(m))[0])}{range_boundaries(str(m))[1]}", 
                    cells_in_range=[f"{get_column_letter(c)}{r}" for r in range(range_boundaries(str(m))[1], range_boundaries(str(m))[3]+1) for c in range(range_boundaries(str(m))[0], range_boundaries(str(m))[2]+1)]
                ) for m in ws.merged_cells.ranges
            ]
            cell_dict = {}
            for row in ws.iter_rows():
                for cell in row:
                    f_c = FillModel(
                        rgb=ExcelObjectEngine._safe_get_rgb(cell.font.color), 
                        theme=cell.font.color.theme if cell.font.color and cell.font.color.type == 'theme' else None, 
                        tint=cell.font.color.tint if cell.font.color and cell.font.color.type == 'theme' else 0.0
                    )
                    fill_m = FillModel(
                        rgb=ExcelObjectEngine._safe_get_rgb(cell.fill.start_color), 
                        theme=cell.fill.start_color.theme if cell.fill.start_color.type == 'theme' else None, 
                        tint=cell.fill.start_color.tint if cell.fill.start_color.type == 'theme' else 0.0
                    )
                    style = StyleModel(
                        font={"name": cell.font.name, "size": cell.font.size, "bold": cell.font.bold, "italic": cell.font.italic, "color": f_c}, 
                        alignment={"horizontal": cell.alignment.horizontal, "vertical": cell.alignment.vertical, "wrap_text": cell.alignment.wrap_text}, 
                        border=BorderModel(
                            left=BorderSideModel(style=cell.border.left.style, color=ExcelObjectEngine._safe_get_rgb(cell.border.left.color)), 
                            right=BorderSideModel(style=cell.border.right.style, color=ExcelObjectEngine._safe_get_rgb(cell.border.right.color)), 
                            top=BorderSideModel(style=cell.border.top.style, color=ExcelObjectEngine._safe_get_rgb(cell.border.top.color)), 
                            bottom=BorderSideModel(style=cell.border.bottom.style, color=ExcelObjectEngine._safe_get_rgb(cell.border.bottom.color))
                        ), 
                        fill=fill_m, number_format=cell.number_format
                    )
                    cell_dict[cell.coordinate] = CellModel(coordinate=cell.coordinate, value=cell.value, style=style)
            
            # [CHECK] row_dimensions와 col_dimensions 추출 시 0 또는 None 값 방어
            row_heights = {int(r): DimensionModel(value=float(rd.height), hidden=bool(rd.hidden)) for r, rd in ws.row_dimensions.items() if rd.height}
            col_widths = {str(c): DimensionModel(value=float(cd.width), hidden=bool(cd.hidden)) for c, cd in ws.column_dimensions.items() if cd.width}
            
            sheet_models.append(SheetModel(sheet_name=ws.title, cells=cell_dict, merged_cells=merged_list, row_heights=row_heights, col_widths=col_widths))
        return WorkbookModel(sheets=sheet_models)

    @staticmethod
    def export_from_model(model: WorkbookModel, output_path: str):
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)
        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        
        for s_model in model.sheets:
            ws = new_wb.create_sheet(title=s_model.sheet_name)
            ws.sheet_view.showGridLines = False
            
            # [1] 너비/높이 레이아웃 우선 적용 (데이터가 채워지기 전에 틀을 먼저 잡음)
            for c_letter, dim in s_model.col_widths.items():
                ws.column_dimensions[str(c_letter)].width = dim.value
                ws.column_dimensions[str(c_letter)].hidden = dim.hidden
            for r_idx, dim in s_model.row_heights.items():
                ws.row_dimensions[int(r_idx)].height = dim.value
                ws.row_dimensions[int(r_idx)].hidden = dim.hidden

            # [2] 전체 배경 화이트 초기화 (범위를 조금 넉넉히 150x60으로 확장)
            for r in range(1, 151):
                for c in range(1, 61): 
                    ws.cell(row=r, column=c).fill = white_fill

            # [3] 셀 데이터 및 스타일 주입
            for coord, c_model in s_model.cells.items():
                cell = ws[coord]
                cell.value = c_model.value
                
                f_cfg = c_model.style.font.copy()
                f_color_val = f_cfg.pop("color", None)
                
                if f_color_val is not None and not isinstance(cell.value, CellRichText):
                    c_obj = ExcelObjectEngine._to_openpyxl_color(f_color_val)
                    cell.font = Font(**f_cfg, color=c_obj) if c_obj else Font(**f_cfg)
                else:
                    cell.font = Font(**f_cfg)
                
                b = c_model.style.border
                cell.border = Border(
                    left=Side(style=b.left.style, color=ExcelObjectEngine._to_openpyxl_color(b.left.color)), 
                    right=Side(style=b.right.style, color=ExcelObjectEngine._to_openpyxl_color(b.right.color)), 
                    top=Side(style=b.top.style, color=ExcelObjectEngine._to_openpyxl_color(b.top.color)), 
                    bottom=Side(style=b.bottom.style, color=ExcelObjectEngine._to_openpyxl_color(b.bottom.color))
                )
                
                fill_obj = ExcelObjectEngine._to_openpyxl_color(c_model.style.fill)
                if fill_obj: 
                    cell.fill = PatternFill(start_color=fill_obj, end_color=fill_obj, fill_type='solid')
                
                cell.alignment = Alignment(**c_model.style.alignment)
                cell.number_format = c_model.style.number_format

            # [4] 병합 및 테두리 보정
            for m_obj in s_model.merged_cells:
                try: ws.merge_cells(m_obj.range_string)
                except: continue
                
                min_c, min_r, max_c, max_r = range_boundaries(m_obj.range_string)
                master = s_model.cells.get(m_obj.master_coordinate)
                if master:
                    b_m = master.style.border
                    for r in range(min_r, max_r + 1):
                        for c in range(min_c, max_c + 1):
                            curr = ws.cell(row=r, column=c)
                            curr.border = Border(
                                left=Side(style=b_m.left.style, color=ExcelObjectEngine._to_openpyxl_color(b_m.left.color)) if c == min_c else curr.border.left,
                                right=Side(style=b_m.right.style, color=ExcelObjectEngine._to_openpyxl_color(b_m.right.color)) if c == max_c else curr.border.right,
                                top=Side(style=b_m.top.style, color=ExcelObjectEngine._to_openpyxl_color(b_m.top.color)) if r == min_r else curr.border.top,
                                bottom=Side(style=b_m.bottom.style, color=ExcelObjectEngine._to_openpyxl_color(b_m.bottom.color)) if r == max_r else curr.border.bottom
                            )

        new_wb.save(output_path)