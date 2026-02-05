import pandas as pd
from copy import deepcopy
from openpyxl.utils import get_column_letter
from src.xl_obj_engine.core.engine import ExcelObjectEngine
from src.xl_obj_engine.core.models import FillModel, DimensionModel # DimensionModel 추가
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
import time


def set_starlet_superscript(sheet_model, coord):
	""" STARLET 문자열에서 LET을 윗첨자로 처리 (Bold 포함) """
	if coord not in sheet_model.cells: return
	superscript_font = InlineFont(vertAlign='superscript', b=True, sz=10) 
	rich_text = CellRichText("STAR", TextBlock(superscript_font, "LET"))
	sheet_model.cells[coord].value = rich_text	
		
	font_style = sheet_model.cells[coord].style.font
	if isinstance(font_style, dict): font_style["color"] = None
	else: font_style.color = None

def set_signature_rich_text(sheet_model, coord, name):
	""" 특정 셀에 '이름(검정) (서명)(회색, Bold)' 부분 서식 적용 """
	if coord not in sheet_model.cells: return
	signature_font = InlineFont(color="AEAAAA", b=True, sz=10)	 
	rich_text = CellRichText(f"{name}", TextBlock(signature_font, "(서명)"))
	sheet_model.cells[coord].value = rich_text
		
	font_style = sheet_model.cells[coord].style.font
	if hasattr(font_style, "get"): font_style["color"] = None
	else: font_style.color = None

def run_report_generation():
	engine = ExcelObjectEngine()
	template_path = "data/templete.xlsx"
	input_path = "data/input_data_sample.xlsx"
	output_path = "output/final_report.xlsx"

	s = time.time()

	# 1. 데이터 로드 및 전처리
	df = pd.read_excel(input_path).ffill()
	df = df.iloc[1:].reset_index(drop=True) 
	
	samples = []
	for i in range(0, len(df), 3):
		chunk = df.iloc[i:i+3]
		if chunk.empty or len(chunk) < 3: break
		samples.append({
			"name": str(chunk.iloc[0]["검체명"]),
			"date": str(chunk.iloc[0]["시험일자"]).split(' ')[0],
			"lot": str(chunk.iloc[0]["Plate Lot No."]),
			"results": chunk.iloc[:, 7:16].values.tolist()
		})

	# 2. 템플릿 로드
	base_template_model = engine.read_to_model(template_path)
	base_sheet = base_template_model.sheets[0]
	final_sheets = []
		
	# 3. 시트별 데이터 주입
	for sheet_idx, i in enumerate(range(0, len(samples), 10)):
		chunk_10 = samples[i:i+10]
		new_sheet = deepcopy(base_sheet)
		new_sheet.sheet_name = f"Rawdata_Page_{chunk_10[0]['date']}_{sheet_idx + 1}"
		
		# 공통 헤더 정보
		new_sheet.cells["K5"].value = chunk_10[0]["date"]
		new_sheet.cells["S6"].value = chunk_10[0]["lot"]

		for idx, sample in enumerate(chunk_10):
			new_sheet.cells[f"E{10+idx}"].value = sample["name"]
			start_row = 45 + (idx * 3) if idx < 5 else 63 + ((idx - 5) * 3)
			
			for r_offset in range(3):
				new_sheet.cells[f"F{start_row + r_offset}"].value = sample["name"]
			
			for r_idx, row_values in enumerate(sample["results"]):
				for c_idx, val in enumerate(row_values):
					col_letter = get_column_letter(7 + c_idx)
					coord = f"{col_letter}{start_row + r_idx}"
					if coord in new_sheet.cells:
						new_sheet.cells[coord].value = val
		
		# [스타일 1] Q43, Q44 빨간색 처리
		for target_coord in ["Q43", "Q44"]:
			if target_coord in new_sheet.cells:
				font_style = new_sheet.cells[target_coord].style.font
				if isinstance(font_style, dict): font_style["color"] = "FF0000"
				else: font_style.color = "FF0000"
		
		# [스타일 2] STARLET 윗첨자 (J22)
		set_starlet_superscript(new_sheet, "J22")

		# [스타일 3] 서명 2인 (C6: 우영진, K6: 문영호)
		set_signature_rich_text(new_sheet, "C6", "				  우영진			  ")
		set_signature_rich_text(new_sheet, "K6", "		  문영호		   ")
		
		# [스타일 4] G-Y 열 너비 5.5 고정 (반드시 append 이전에 실행)
		for col_idx in range(7, 26): 
			col_letter = get_column_letter(col_idx)
			new_sheet.col_widths[col_letter] = DimensionModel(value=6.07, hidden=False)

		# 모든 스타일이 반영된 시트를 리스트에 추가
		final_sheets.append(new_sheet)

	# 4. 저장
	base_template_model.sheets = final_sheets
	engine.export_from_model(base_template_model, output_path)
	print(f"Report Generated at {output_path}")
	e = time.time()
	
	print(f"Elapsed Time: {e - s:.2f} seconds")

if __name__ == "__main__":
	run_report_generation()