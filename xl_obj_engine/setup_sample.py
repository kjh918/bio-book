import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Analysis_Report"

# 레이아웃 설정
ws.column_dimensions['A'].width = 30
ws.row_dimensions[1].height = 30

# 병합 및 테두리
ws.merge_cells("A1:C1")
cell = ws["A1"]
cell.value = "BIOINFORMATICS PIPELINE RESULT"
cell.font = Font(bold=True, size=14)
cell.alignment = Alignment(horizontal="center", vertical="center")
thin = Side(style='thin')
cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

wb.save("data/input_template.xlsx")
print("Sample template created in 'data/'")
